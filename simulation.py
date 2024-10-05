# -----------------------------------------------------------------------------
# Author: Ludwig Horvath
# Email: ludhor@kth.se
# Date: 2024-10-05
# -----------------------------------------------------------------------------


# Standard library imports
import os
import random

# Numerical and data handling libraries
import numpy as np
import pandas as pd
from scipy.integrate import quad
import matplotlib.pyplot as plt

# Third-party libraries
from openpyxl import load_workbook
from tqdm import tqdm

# Custom modules
from settings import saved_features

# Global constants
dt = 0.1  # Time step for the simulation; smaller values provide better accuracy

# ===================================
# Classes
# ===================================


class Sensor:
    def __init__(self, P_u, P_f, std):
        """
        Initialize the sensor with given probabilities for its components and accuracy.
        P_u: Probability of selecting uniform distribution.
        P_n: Probability of selecting normal distribution.
        P_f: Probability of selecting failure distribution.
        std: Standard deviation for the normal distribution.
        """
        self.P_u = P_u
        self.P_f = P_f
        self.P_n = 1 - P_u - P_f
        
        self.Dran = 10
        self.D_fail = self.Dran + 0.2
        self.std = std

        self.d = None

    def sample(self, d):
        self.d = d
        """Sample a value based on the distribution probabilities."""
        rand_val = np.random.rand()  # Random value in [0, 1]

        if rand_val < self.P_u:  # Uniform distribution
            return self.sample_uniform()
        elif rand_val < self.P_u + self.P_n:  # Normal distribution
            return self.sample_normal()
        else:  # Failure distribution
            return self.sample_failure()

    def sample_uniform(self):
        """Sample from uniform distribution [0, 10]."""
        return np.random.uniform(0, self.Dran)

    def sample_normal(self):
        """Sample from normal distribution with specified accuracy."""
        return np.random.normal(self.d, self.std*self.d)

    def sample_failure(self):
        """Sample from uniform distribution with a shift to ensure values > 10."""
        return np.random.uniform(self.Dran, self.D_fail)

    def fu(self, x):
        """Uniform distribution component (for reference)."""
        if 0 <= x <= self.Dran:
            return self.P_u / self.Dran
        else:
            return 0

    def fn(self, x):
        """Normal distribution component (for reference)."""
        return (1 / (np.sqrt(2 * np.pi) * self.std)) * \
               np.exp(-(x - self.d) ** 2 / (2 * self.std ** 2))

    def fb(self, x):
        """Binary/Uniform distribution component (for reference)."""
        if self.Dran < x <= (self.Dran + self.D_fail):
            return self.P_f / self.D_fail
        else:
            return 0
        
    def sample_distribution(self, num_samples=1000):
        """
        Sample from the combined distribution using rejection sampling.
        num_samples: Number of samples to generate.
        """

        self.d = float(input('Input True Distance "d" in [0, 10]: '))

        samples = []
        
        # Calculate the maximum value of the envelope function

        x = np.random.uniform(0, self.D_fail + self.Dran)  # Sample within the total range

        max_f = max(self.P_u / self.Dran, 
                     self.P_f / self.D_fail, 
                     self.fn(x) + 0.1)  # Slightly adjust for normal envelope
        
        while len(samples) < num_samples:
            y = np.random.uniform(0, max_f)  # Sample from the envelope height
            
            # Check if we accept the sample based on the combined pdf
            if (y <= self.fu(x)) or (y <= self.fn(x)) or (y <= self.fb(x)):
                samples.append(x)

        return samples

    def plot_distribution(self, num_samples=100, bins=10):
        """
        Plot the distribution of sampled values.
        num_samples: Number of samples to generate for the plot.
        bins: Number of bins for the histogram.
        """
        samples = self.sample_distribution(num_samples)
        
        plt.figure(figsize=(10, 6))
        plt.hist(samples, bins=bins, density=True, alpha=0.6, color='g', label='Sampled Distribution')

        # Plot the theoretical distributions for comparison
        x_values = np.linspace(-1, self.D_fail + self.Dran + 1, 1000)
        plt.plot(x_values, [self.fu(x) for x in x_values], 'r-', label='Uniform Distribution (f_u)', linewidth=2)
        plt.plot(x_values, [self.fn(x) for x in x_values], 'b-', label='Normal Distribution (f_n)', linewidth=2)
        plt.plot(x_values, [self.fb(x) for x in x_values], 'c-', label='Failure Distribution (f_f)', linewidth=2)

        plt.title('Distribution of Samples from Sensor')
        plt.xlabel('Value')
        plt.ylabel('Density')
        plt.legend()
        plt.grid()
        plt.show()


class Robot:
    def __init__(self, sensor):
        """
        Initialize the Robot with a given sensor, mode, and movement parameters.
        """
        self.sensor = sensor  # Detection system
        self.mode = False  # Mode of robot (Radiation + Driving or Driving)
        self.uv = True  # UV-light status (True if UV light is on)
        self.alarm = False  # Alarm status (True if alarm is active)
        self.v = 2/3.6  # Speed of the robot
        self.d_star = 100  # Distance to the closest human
        self.position = None  # Robot's current position
        self.velocity = np.array([0, 0])  # Current velocity vector

        self.setpoints = {'coord': [], 'mode': []}  # Path setpoints
        self.setpoint_nr = 1  # Current setpoint index
        self.setpoint_goal = None  # Final setpoint index
        self.at_goal = False  # Flag indicating if the goal is reached

        self.tolerance = 0.3  # Position error tolerance

        # Initialize the path for the robot based on saved features
        self.init_path(saved_features)

    def update_measurement(self, d_truth):
        """
        Update the robot's measurement of the distance to the closest human using the sensor.
        """
        d_star = np.min(self.sensor.sample(d_truth))  # Sample from the sensor
        thresholds = np.array([3, 7, 10, 100])
        # Update d_star based on thresholds
        self.d_star = np.min(thresholds[thresholds > d_star]) if np.any(thresholds > d_star) else None

    def init_path(self, features):
        """
        Initialize the robot's path and setpoints based on the provided path features.
        """
        for path in features['paths']:
            # Get x, y coordinates from the path data
            x_data = features['paths'][path]['Line'].get_xdata()
            y_data = features['paths'][path]['Line'].get_ydata()
            xy_data = np.column_stack((x_data, y_data))  # Pair x and y coordinates
            self.setpoints['coord'].extend(xy_data.tolist())  # Add coordinates to setpoints

            # Extend the mode for each point in the path
            n = len(x_data)
            self.setpoints['mode'].extend(features['paths'][path]['M'] for _ in range(n))

        self.setpoint_goal = len(self.setpoints['coord']) - 1  # Set goal index
        self.position = np.array(self.setpoints['coord'][0])  # Set initial position
        self.setpoint = np.array(self.setpoints['coord'][self.setpoint_nr])  # Set initial target

    def update_error(self):
        """
        Update the error between the robot's position and the current setpoint.
        """
        self.error = np.linalg.norm(self.setpoint - self.position)  # Euclidean distance to setpoint

    def update_setpoint(self):
        """
        Update the current setpoint if the robot is close enough to the current target.
        """
        if self.error <= self.tolerance:
            if self.setpoint_nr == self.setpoint_goal:
                self.at_goal = True  # Goal reached
            else:
                self.setpoint_nr += 1  # Move to the next setpoint
                self.setpoint = np.array(self.setpoints['coord'][self.setpoint_nr])

    def update_state_machine(self):
        """
        Update the robot's state based on its current position, setpoint, and human proximity.
        """
        self.update_error()  # Calculate error to setpoint
        self.update_setpoint()  # Update setpoint if needed

        # Reset alarm and UV status
        self.alarm = False
        self.uv = True

        # Calculate new velocity based on residual distance to the setpoint
        residual = self.setpoint - self.position
        self.velocity = np.round(self.v * residual / np.linalg.norm(residual), 2)

        self.mode = self.setpoints['mode'][self.setpoint_nr]

        # Check for human proximity and react accordingly
        if self.d_star <= 10:
            self.alarm = True
            if self.d_star <= 7:
                if self.mode == 1:
                    self.velocity = np.array([0, 0])  # Stop the robot
                self.uv = False  # Turn off UV light
                if self.d_star <= 3:
                    pass  # Further actions if needed

        # Override UV light if in the wrong mode
        if not self.mode:
            self.uv = False


class Human:
    def __init__(self, trained):
        """
        Initialize a Human with random characteristics and setpoint behavior.
        """
        self.debug = False

        self.id = np.random.randint(0, 100000)  # Unique identifier
        self.trained = trained  # Is the human trained?

        if self.trained:
            self.p = saved_features['agents']['human']['worker']['p_react_to_alarm']
        else:
            self.p = saved_features['agents']['human']['visitor']['p_react_to_alarm']

        self.alert = np.random.binomial(n=1, p=self.p)
        self.v_walk = 5 / 3.6  # Walking speed in m/s
        self.v_run = 13 / 3.6  # Running speed in m/s
        self.lambda_P = 1
        self.lambda_D = 0.1
        self.f_t_old = 0
        self.alpha = 100
        self.d_threshold = -10 * self.alpha  # Distance threshold for scaling

        self.spawned = False  # Is the human in the simulation?

        # Movement and position attributes
        self.leaving = None
        self.chasing = None
        self.setpoint = None
        self.error = None
        self.tolerance = 1
        self.position = None
        self.velocity = None

        self.checkpoints = saved_features['checkpoints']  # Checkpoints to navigate
        features = saved_features['agents']['human']['worker'] if self.trained else saved_features['agents']['human']['visitor']
        self.P = np.array(features['P'])  # Transition matrix for checkpoint navigation
        self.color = features['color']  # Human's display color

    def update_error(self):
        """
        Update the current error between the human and the setpoint.
        """
        self.error = np.linalg.norm(self.setpoint - self.position)

    def update_setpoint(self):
        """
        Update the setpoint when the human reaches the current setpoint.
        """
        if self.error <= self.tolerance:
            # Set new leaving and chasing checkpoints
            self.leaving['label'] = self.chasing['label']
            self.leaving['coordinate'] = self.chasing['coordinate']

            checkpoints_list = list(self.checkpoints.keys())
            i = checkpoints_list.index(self.leaving['label'])
            weights = self.P[i, :]
            j = random.choices(range(len(checkpoints_list)), weights=weights)[0]

            chasing_label = checkpoints_list[j]
            chasing_values = self.checkpoints[chasing_label]
            chasing_coordinate = np.array([chasing_values['x'], chasing_values['y']])
            noise = np.random.normal(0, 0.1, chasing_coordinate.shape)  # Add random noise
            chasing_coordinate += noise

            self.chasing = {'label': chasing_label, 'coordinate': chasing_coordinate}
            distance_node2node = np.linalg.norm(chasing_coordinate - self.leaving['coordinate'])
            walk_gradient = (chasing_coordinate - self.leaving['coordinate']) / distance_node2node
            self.velocity = walk_gradient * self.v_walk
            self.setpoint = chasing_coordinate

    def spawn(self, anywhere=True):
        # Spawn a human at a specified location, optionally forcing initial spawn to an entrance.
        if anywhere:
            init_checkpoint_pair = dict(random.sample(list(self.checkpoints.items()), 2))  # Randomly select two checkpoints
        else:
            # Ensure the human spawns at 'entrance/exit' checkpoint and another random checkpoint
            available_checkpoints = {k: v for k, v in self.checkpoints.items() if k != 'entrance/exit'}
            init_checkpoint_pair = {'entrance/exit': self.checkpoints['entrance/exit'],
                                    **random.sample(list(available_checkpoints.items()), 1)[0]}

        # Extract coordinates of the selected checkpoints
        node_pair_keys = list(init_checkpoint_pair.keys())
        node_pair_values = list(init_checkpoint_pair.values())
        node_1_coord = np.array([node_pair_values[0]['x'], node_pair_values[0]['y']])
        node_2_coord = np.array([node_pair_values[1]['x'], node_pair_values[1]['y']])

        # Set initial leaving and chasing coordinates
        self.leaving = {'label': node_pair_keys[0], 'coordinate': node_1_coord}
        self.chasing = {'label': node_pair_keys[1], 'coordinate': node_2_coord}

        self.setpoint = node_2_coord  # Target coordinate
        
        distance_node2node = np.linalg.norm(node_2_coord - node_1_coord, 2)  # Calculate distance between checkpoints
        walk_gradient = (node_2_coord - node_1_coord) / distance_node2node  # Determine direction of movement
        d_rand = random.uniform(0, distance_node2node)  # Random distance to spawn within the node
        self.position = np.array(node_1_coord + walk_gradient * d_rand).reshape(1,2)
        self.velocity = walk_gradient*self.v_walk  # Set initial velocity based on direction

        self.spawned = True  # Mark the human as spawned


    def update_state_machine(self, robot):
        # Update the human's state based on their current status and position relative to the robot.
        if self.spawned:
            self.update_error()  # Calculate current error to the setpoint
            self.update_setpoint()  # Update the setpoint if within tolerance

            # Directly set the velocity towards the setpoint
            direction_to_setpoint = self.setpoint - self.position
            direction_to_setpoint_norm = np.linalg.norm(direction_to_setpoint)

            if direction_to_setpoint_norm > 0:  # Avoid division by zero
                
                r_l = self.position
                r_c = self.chasing['coordinate']     

                v_N_vec = (r_c  - r_l)/(np.linalg.norm(r_c - r_l))*self.v_walk
                

                if self.alert:
                    d_vec = self.position - robot.position
                    d = np.linalg.norm(d_vec)
                    f_t = 1/d**2 * d_vec
                    df_t = f_t - self.f_t_old
                    self.f_t_old = f_t
                    scale_factor = 1 - 1/(1 + np.exp(-(self.alpha*d + self.alpha * self.d_threshold)))
                    v_R_vec = (self.lambda_P * f_t + self.lambda_D * df_t / dt)*scale_factor
                    
                else:
                    v_R_vec = 0


                v_vec = v_R_vec + v_N_vec

                v = np.linalg.norm(v_vec)

                self.velocity = v_vec / v * np.min([v, self.v_run])
                
                if self.debug: print(f'ID: {self.id}, v_N_vec: {v_N_vec}, d: {d}, v_R_vec: {v_R_vec}, scale_factor: {scale_factor}, final_velocity: {self.velocity}')
                if self.debug: input()

        else:
            print('Human not spawned')  # Indicate that the human is not active

class Env:
    def __init__(self, robot, nr_workers, nr_visitors):
        """
        Initialize the Environment with its features and agents.
        """
        self.robot = robot
        self.state = None

        # Define the origin of the environment
        self.origin = [0, 0]  # Origin point
        
        # Initialize a dictionary to store all agents (robot, workers, visitors)
        self.agents = {}

        self.nr_workers = nr_workers
        self.nr_visitors = nr_visitors

    def init_robot_workers_visitors(self):
        """
        Initialize and spawn the robot, workers, and visitors into the environment.
        """
        # Add the robot to the agents dictionary
        self.agents['robot'] = self.robot
        
        # Add the workers to the environment
        for nr in range(1, self.nr_workers + 1):
            worker = Human(trained=True)
            worker.spawn(anywhere=True)
            self.agents[f'worker{nr}'] = worker

        # Add the visitors to the environment
        for nr in range(1, self.nr_visitors + 1):
            visitor = Human(trained=False)
            visitor.spawn(anywhere=True)
            self.agents[f'visitor{nr}'] = visitor

    def get_state(self):
        """
        Get the current state of the environment, including the robot and human statuses.
        """
        # Get UV status and position of the robot
        U = self.agents['robot'].uv  # Robot's UV status
        robot_position = self.agents['robot'].position  # Robot's current position

        # Get robot mode
        M = self.agents['robot'].mode

        # List to hold distances between the robot and humans
        d_list = []

        # Calculate distances between the robot and all humans in the environment
        for agent in self.agents.values():
            if isinstance(agent, Human):
                d_list.append(np.linalg.norm(agent.position - robot_position, 2))
        
        # Find the minimum distance to the closest human
        d = np.min(d_list)

        # Update the robot's measurement for the closest human distance
        self.agents['robot'].update_measurement(d)
        d_star = self.agents['robot'].d_star  # Get robot's D* value
         
        # Determine the appropriate distance threshold (D)
        thresholds = np.array([3, 7, 10, 100])       
        d = np.min(thresholds[thresholds > d]) if np.any(thresholds > d) else None

        # Return the current state as a dictionary
        return {'M': M, 'U': U, 'D*': d_star, 'D': d}

    def get_positions(self):
        """
        Get the positions of all agents in the environment.
        """
        # Dictionary to store agent positions
        position_dict = {}

        # Retrieve and store positions for each agent
        for agent_name in self.agents.keys():
            position_dict[agent_name] = self.agents[agent_name].position

        return position_dict

    def remove_human(self, human_key):
        """
        Remove a human from the environment.
        """
        # Set the specified human agent to None, effectively removing them
        self.agents[human_key] = None  

    def propagate(self, dt):
        """
        Update the state of all agents and refresh dynamic features over the time step dt.
        """
        # Iterate through all agents in the environment
        for agent_key in self.agents.keys():
            agent = self.agents[agent_key]  # Retrieve the current agent

            # Update the agent's state machine (differentiates between human and robot)
            if isinstance(agent, Human):
                agent.update_state_machine(self.robot)
            else:
                agent.update_state_machine()

            # Update the agent's position based on its velocity and the time step
            agent.position = agent.position + agent.velocity * dt


class Sim:
    def __init__(self, sensor_parameters, nr_workers, nr_visitors, runs=100):
        """
        Initialize the simulation with a robot, environment, and configuration parameters.
        """
        # Metadata and Debugging
        self.metadata = {'Iteration': [], 'Outcome': [], 'Duration': []}
        self.debug = False
        self.time = 0  # Current simulation time
        self.runs = runs  # Number of simulation runs
        
        # Sensor and Environment
        self.nr_workers = nr_workers
        self.nr_visitors = nr_visitors
        self.sensor_args = sensor_parameters 
        
        # State Management
        no_states = 2*2*4*4+2  # Number of states in the state space
        self.p_matrix_evolution = np.zeros((no_states, no_states, self.runs))  # Placeholder for state transition matrix
        
        # Data Storage for Results
        self.data = {'time': [], 'position': [], 'state': []}
        self.dataG = None
        self.dataI = None
        self.first_G = None
        self.first_I = None

        # Average Matrices
        self.P_avg = None
        self.B_avg = None

    def store_metadata(self, iteration, outcome):
        """
        Store metadata for a specific run of the simulation.
        """
        self.metadata['Iteration'].append(iteration)
        self.metadata['Outcome'].append(outcome)
        self.metadata['Duration'].append(self.time)

    def reset_data_placeholder(self):
        """
        Reset data storage to an empty placeholder after each run.
        """
        self.data = {'time': [], 'position': [], 'state': []}

    def sample_transition(self, state_i, state_j, iteration, robot):
        """
        Sample and update the transition matrix based on state changes.
        """
        M = [0, 1]
        U = [0, 1]
        D = [100, 10, 7, 3]
        D_star = [100, 10, 7, 3]

        i = 32 * M.index(state_i['M']) + 16 * U.index(state_i['U']) + 4 * D.index(state_i['D*']) + D_star.index(state_i['D'])
        j = 32 * M.index(state_j['M']) + 16 * U.index(state_j['U']) + 4 * D.index(state_j['D*']) + D_star.index(state_j['D'])

        outcome = None

        if robot.at_goal:
            j = 64
            self.absorbed = True
            if self.first_G is None:
                self.first_G = iteration
            if self.debug: input(f'Goal State Reached {state_j}')
            outcome = 'G'

        elif state_i['M'] == 1:
            if state_i['U'] == 1:
                if state_i['D'] <= 3:
                    j = 65
                    self.absorbed = True
                    if self.first_I is None:
                        self.first_I = iteration
                    if self.debug: input(f'Injury State Reached {state_j}')
                    outcome = 'I'

        self.p_matrix_evolution[i, j, iteration] += 1
        return outcome

    def run(self):
        """
        Run the simulation for a defined number of iterations.
        """
        last_state = None

        for iteration in tqdm(range(self.runs)):  # Loop over the number of runs
            sensor = Sensor(**self.sensor_args)
            robot = Robot(sensor)
            environment = Env(robot, nr_workers=self.nr_workers, nr_visitors=self.nr_visitors)
            environment.init_robot_workers_visitors()
            self.time = 0  # Current simulation time
            self.absorbed = False

            while not self.absorbed:  # Continue until the environment is absorbed
                environment.propagate(dt)  # Update environment state
                state = environment.get_state()  # Get the current state
                positions = environment.get_positions()  # Get the current positions

                if last_state is None:
                    last_state = state
                else:
                    outcome = self.sample_transition(last_state, state, iteration, robot)
                    last_state = state

                if not (self.first_G or self.first_I):
                    self.data['time'].append(self.time)
                    self.data['position'].append(positions)
                    self.data['state'].append(state)

                self.time += dt  # Increment the time by the time step

            self.store_metadata(iteration, outcome)

            if iteration == self.first_G:
                self.dataG = self.data
                self.reset_data_placeholder()

            if iteration == self.first_I:
                self.dataI = self.data
                self.reset_data_placeholder()

    def calc_transition_probabilities(self):
        """
        Calculate the average transition matrix from simulation data.
        """
        self.P_avg = np.sum(self.p_matrix_evolution, axis=2)
        
        row_sums = np.sum(self.P_avg, axis=1)

        self.P_avg_normalized = self.P_avg / row_sums[:, np.newaxis]

        # Normalize the rows
        row_sums = self.P_avg.sum(axis=1, keepdims=True)
        row_sums[row_sums == 0] = 1  # Avoid division by zero
        self.P_avg = self.P_avg / row_sums

        # Absorbing states
        self.P_avg[64, 64] = 1
        self.P_avg[65, 65] = 1

        return self.P_avg

    def calc_absorption_probabilities(self):
        """
        Calculate the absorbing probabilities of the Markov chain.
        """
        if self.P_avg is None:
            print("You must calculate the transition matrix first.")
            return None

        # Extract Q and R
        n_transient = 64
        Q = self.P_avg[:n_transient, :n_transient]
        R = self.P_avg[:n_transient, n_transient:]

        # Compute the fundamental matrix N
        I = np.eye(n_transient)
        I_minus_Q = I - Q
        try:
            N = np.linalg.inv(I_minus_Q)
        except np.linalg.LinAlgError:
            print("Warning: The matrix (I - Q) is singular.")
            N = np.linalg.pinv(I_minus_Q)

        # Calculate absorbing probabilities
        self.B_avg = N @ R
        return self.B_avg

    def calc_performance_metrics(self, data):
        """
        Calculates performance metrics based on the provided DataFrame.
        """
        print(data)


        total_iterations = self.runs
        num_injuries = (data['Outcome'] == 'I').sum()
        risk_of_injury = num_injuries / total_iterations if total_iterations > 0 else 0
  
        performance_metrics = {
            'N': total_iterations,
            '# I': num_injuries,
            'Risk': risk_of_injury,
        }

        return performance_metrics

    def save_calc_model(self):
        """
        Save both the P matrix and B matrix to an Excel file, along with performance metrics.
        """
        base_directory = 'Results'
        today = pd.Timestamp.now().strftime('%Y-%m-%d')
        time_of_day = pd.Timestamp.now().strftime('%H-%M')
        folder_path = os.path.join(base_directory, today, time_of_day)
        os.makedirs(folder_path, exist_ok=True)


        self.calc_transition_probabilities()
        # Save the P matrix to a DataFrame
        df_P = pd.DataFrame(self.P_avg)

        # Load the existing Excel file
        old_file_path = os.path.join(base_directory, 'EmptyTransitionMatrix.xlsx')
        wb = load_workbook(old_file_path)
        ws = wb['TransitionMatrix']

        # Write the P matrix to the sheet
        start_row_P = 1
        start_col_P = 1
        for i in range(df_P.shape[0]):
            for j in range(df_P.shape[1]):
                ws.cell(row=start_row_P + i, column=start_col_P + j, value=df_P.iat[i, j])

        # Write the B matrix values if available
        self.calc_absorption_probabilities()
        if self.B_avg is not None:
            df_B = pd.DataFrame(self.B_avg, columns=['Absorbing State 65', 'Absorbing State 66'])
            start_row_B = 74
            for j in range(min(df_B.shape[0], 64)):
                ws.cell(row=start_row_B, column=j + 1, value=df_B.iat[j, 0])
                ws.cell(row=start_row_B + 1, column=j + 1, value=df_B.iat[j, 1])

        # Save performance metrics
        performance_metrics = self.calc_performance_metrics(pd.DataFrame(self.metadata))
        start_row_M = 74
        start_col_M = 67
        for j, key in enumerate(performance_metrics.keys()):
            ws.cell(row=start_row_M + 1, column=start_col_M + j + 1, value=performance_metrics[key])

        # Save the workbook to the new folder
        new_file_path = os.path.join(folder_path, f'P_and_B_{time_of_day}.xlsx')
        wb.save(new_file_path)

        # Open the Excel file automatically
        os.startfile(new_file_path)

    